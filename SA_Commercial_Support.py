# coding=utf-8
import os

import xlrd
import logging
from openpyxl import load_workbook
from copy import copy

logging.basicConfig(filename='mylog.txt', format="%(asctime)s : %(message)s",
                    level=logging.DEBUG)
# 读数据时，计算各指标值
sum_target_col_list = {'日期': '', '利用率平均值': 0, 'Fallback流程成功率': 0, '发送PDU Session Establishment Accept消息数': 0,
                       '接收PDU Session Establishment Request消息数': 0,
                       'N2模式二次寻呼响应次数': 0, 'N2模式寻呼请求次数': 0, 'N2模式一次寻呼响应次数': 0,
                       'AMF注册态最大用户数': 0, 'PGW-C通过GTP S5/S8/S2a/S2b创建的平均在线Session数': 0, 'SMF 5G 实时PDU会话数': 0, }

# 获取到各指标所在列（value），用于写数据时定位到具体cell
data_target_col_list = {'日期': '', '今日活动用户数(5G)': 0, '今日活动用户数(5G+4G)': 0, '昨日活动用户数(5G)': 0, '昨日活动用户数(5G+4G)': 0,
                        '新增日活数(5G)': 0, '新增日活数(5G+4G)': 0, 'SA寻呼成功率': 0, 'SA会话建立成功率': 0, 'EPS FB接通率': 0,
                        'AMF CPU负荷': 0, 'SMF CPU负荷': 0}

VMDic = {'AMF_CPU_Rate': 0, 'AMF_Count': 0, 'SMF_CPU_Rate': 0, 'SMF_Count': 0}


# 获取当前路径下所有xls和xlsx文件
def XLSFileList():
    filelist = []
    for root, dirs, files in os.walk(".", topdown=False):
        for name in files:
            str = os.path.join(root, name)
            if str.split('.')[-1] == 'xls' or str.split('.')[-1] == 'xlsx':
                filelist.append(str)
    return filelist


# 读取配置文件
def readConfig():
    try:
        confDic = {}
        file = open(os.getcwd() + '\\config.txt', "rb")
        for line in file.readlines():
            if 'region' in line.decode():
                confDic['region'] = line.decode().split('=')[1].strip('\n\t\r')
    except Exception as err:
        confDic['region'] = '广东'
        logging.error('readConfig function error : %s', err)
    return confDic


# 计算总和
def MaxNum(rown, coln, num_rows, worksheet):
    datalist = []
    for r in range(rown, num_rows):
        try:
            datalist.append(int(worksheet.cell_value(r, coln)))
        except Exception as err:
            logging.error('MaxNum function error : %s', err)
            continue
    return max(datalist)


# 计算平均数
def AverageNum(rown, coln, num_rows, worksheet):
    total = 0
    num = 0
    res = 0
    for r in range(rown, num_rows):
        try:
            total += int(worksheet.cell_value(r, coln))
            num = num + 1
        except Exception as err:
            logging.error('addTotal function error : %s', err)
            continue
    if num != 0:
        res = total / num
    return res


# 计算总和
def TotalNum(rown, coln, num_rows, worksheet):
    total = 0
    for r in range(rown, num_rows):
        try:
            total += int(worksheet.cell_value(r, coln))
        except Exception as err:
            logging.error('addTotal function error : %s', err)
            continue
    return total


# 计算平均值（取除0值）
def AverageNumWithoutZero(rown, coln, num_rows, worksheet):
    total = 0
    num = 0
    res = 0
    for r in range(rown, num_rows):
        try:
            if float(worksheet.cell_value(r, coln)) != 0:
                total += float(worksheet.cell_value(r, coln))
                num = num + 1
        except Exception as err:
            logging.error('addTotal function error : %s', err)
            continue
    if num != 0:
        res = total / num
    return res


# 读取模板文件，获取到要填的单元格，将计算的数据写入到目标单元格
def ReadAndAnalysis(XLSPath):
    logging.info('analytical %s start.', XLSPath)
    workbook = xlrd.open_workbook(XLSPath)  # 打开一个workbook
    worksheet = workbook.sheet_by_name(u'Sheet1')  # 定位到目标sheet
    num_rows = worksheet.nrows  # 获取该sheet中的有效行数
    num_cols = worksheet.ncols  # 获取列表的有效列数
    logging.info('begin analytical %s.', XLSPath)
    if '保存时间' in str(worksheet.cell_value(1, 0)):
        sum_target_col_list['日期'] = '.'.join(str(worksheet.cell_value(1, 0)).split(' ')[1].split('-')[1:3])

    # 解析xls
    row = 0
    col = 0
    flag = 0
    Nrow = 0
    Ncol = 0
    Nflag = 0
    title = ''
    for rown in range(0, num_rows):
        for coln in range(num_cols):
            # 解析指定列
            for tclo in sum_target_col_list:
                if tclo in str(worksheet.cell_value(rown, coln)) and coln > 2:
                    if 'AMF注册态最大用户数' == tclo or '实时PDU会话数' in tclo or '创建的平均在线Session数' in tclo:
                        Nrow = rown + 1
                        UserNumList = []
                        while Nrow < num_rows:
                            UserNumList.append(
                                int(worksheet.cell_value(Nrow, coln)) + int(worksheet.cell_value(Nrow + 1, coln)))
                            Nrow = Nrow + 2
                        if len(UserNumList):
                            UserNumList.sort(reverse=True)
                            sum_target_col_list[tclo] = UserNumList[0]
                        break
                    if 'N2模式二次寻呼响应次数' in tclo or 'N2模式寻呼请求次数' in tclo or 'N2模式一次寻呼响应次数' in tclo:
                        sum_target_col_list[tclo] = AverageNum(rown + 1, coln, num_rows, worksheet)
                        break
                    if 'PDU Session Establishment' in tclo:
                        sum_target_col_list[tclo] = TotalNum(rown + 1, coln, num_rows, worksheet)
                        break
                    if 'Fallback流程成功率' in tclo:
                        sum_target_col_list[tclo] = AverageNumWithoutZero(rown + 1, coln, num_rows, worksheet)
                        break
                    if '利用率平均值' in tclo:
                        flag = 1
                        row = rown + 1
                        col = coln
                        break
    if flag == 1:
        for r in range(row, num_rows):
            if 'SMF' in str(worksheet.cell_value(r, col - 2)):
                VMDic['SMF_CPU_Rate'] = VMDic['SMF_CPU_Rate'] + float(worksheet.cell_value(r, col))
                VMDic['SMF_Count'] = VMDic['SMF_Count'] + 1
            if 'AMF' in str(worksheet.cell_value(r, col - 2)):
                VMDic['AMF_CPU_Rate'] = VMDic['AMF_CPU_Rate'] + float(worksheet.cell_value(r, col))
                VMDic['AMF_Count'] = VMDic['AMF_Count'] + 1
        # print(UserNumList)
    logging.info('analytical complete.', )


def ReadAndWrite(XLSPath):
    # 打开一个workbook
    workbook = xlrd.open_workbook(XLSPath)
    # 抓取所有sheet页的名称
    worksheets = workbook.sheet_names()
    # 定位到目标sheet
    worksheet = workbook.sheet_by_name(worksheets[0])
    # 获取该sheet中的有效行数
    num_rows = worksheet.nrows
    # 获取列表的有效列数
    num_cols = worksheet.ncols

    # 打开excel
    wb = load_workbook(XLSPath)
    # 获取工作表
    for name in wb.sheetnames:
        if 'Sheet1' in name:
            sheet = wb[name]

    logging.info('write %s start.', XLSPath)
    for row in range(num_rows):
        CopyCellFont(sheet, row + 1, num_cols + 1)
    sheet.cell(1, num_cols + 1).value = data_target_col_list['日期']
    sheet.cell(2, num_cols + 1).value = data_target_col_list['今日活动用户数(5G)']
    sheet.cell(3, num_cols + 1).value = data_target_col_list['今日活动用户数(5G+4G)']
    sheet.cell(4, num_cols + 1).value = sheet.cell(2, num_cols).value
    sheet.cell(5, num_cols + 1).value = sheet.cell(3, num_cols).value
    sheet.cell(6, num_cols + 1).value = str(data_target_col_list['今日活动用户数(5G)'] - sheet.cell(2, num_cols).value)
    sheet.cell(7, num_cols + 1).value = str(data_target_col_list['今日活动用户数(5G+4G)'] - sheet.cell(3, num_cols).value)
    if data_target_col_list['SA寻呼成功率'] != 0:
        sheet.cell(8, num_cols + 1).value = data_target_col_list['SA寻呼成功率']
    if data_target_col_list['SA会话建立成功率'] != 0:
        sheet.cell(9, num_cols + 1).value = data_target_col_list['SA会话建立成功率']
    sheet.cell(10, num_cols + 1).value = '*(' + str(data_target_col_list['EPS FB接通率']) + ')'
    sheet.cell(11, num_cols + 1).value = str(data_target_col_list['AMF CPU负荷']) + '/' + str(
        data_target_col_list['SMF CPU负荷'])

    wb.save(XLSPath)
    logging.info('write %s end.', XLSPath)


def CopyCellFont(sheet, writeRow, writeCol):
    sheet.cell(writeRow, writeCol).font = copy(sheet.cell(writeRow, 6).font)
    sheet.cell(writeRow, writeCol).border = copy(sheet.cell(writeRow, 6).border)
    sheet.cell(writeRow, writeCol).fill = copy(sheet.cell(writeRow, 6).fill)
    sheet.cell(writeRow, writeCol).number_format = copy(sheet.cell(writeRow, 6).number_format)
    sheet.cell(writeRow, writeCol).protection = copy(sheet.cell(writeRow, 6).protection)
    sheet.cell(writeRow, writeCol).alignment = copy(sheet.cell(writeRow, 6).alignment)


def XLSReadAndAnalysis():
    logging.info('begin to read file')
    for xfl in XLSFileList():
        ReadAndAnalysis(xfl)
    # for i in sum_target_col_list:
    #     print(i, sum_target_col_list[i])


def XLSReadAndWrite():
    logging.info('begin to write file')
    for xfl in XLSFileList():
        if '指标汇总' in xfl:
            ReadAndWrite(xfl)


def ConvertData():
    logging.info('begin to ConvertData')
    data_target_col_list['日期'] = sum_target_col_list['日期']
    data_target_col_list['今日活动用户数(5G)'] = sum_target_col_list['AMF注册态最大用户数']
    data_target_col_list['今日活动用户数(5G+4G)'] = sum_target_col_list['PGW-C通过GTP S5/S8/S2a/S2b创建的平均在线Session数'] + \
                                             sum_target_col_list['SMF 5G 实时PDU会话数']
    if sum_target_col_list['N2模式寻呼请求次数'] != 0:
        data_target_col_list['SA寻呼成功率'] = str(round(
            (sum_target_col_list['N2模式二次寻呼响应次数'] + sum_target_col_list['N2模式一次寻呼响应次数']) / \
            sum_target_col_list['N2模式寻呼请求次数'] * 100, 2)) + '%'
    if sum_target_col_list['接收PDU Session Establishment Request消息数'] != 0:
        data_target_col_list['SA会话建立成功率'] = str(round(
            sum_target_col_list['发送PDU Session Establishment Accept消息数'] / sum_target_col_list[
                '接收PDU Session Establishment Request消息数'] * 100, 2)) + '%'
    data_target_col_list['EPS FB接通率'] = str(round(sum_target_col_list['Fallback流程成功率'], 2)) + '%'
    if VMDic['AMF_Count'] != 0:
        data_target_col_list['AMF CPU负荷'] = str(round(VMDic['AMF_CPU_Rate'] / VMDic['AMF_Count'], 2)) + '%'
    if VMDic['SMF_Count'] != 0:
        data_target_col_list['SMF CPU负荷'] = str(round(VMDic['SMF_CPU_Rate'] / VMDic['SMF_Count'], 2)) + '%'
    # for i in data_target_col_list:
    #     print(i, ' ', data_target_col_list[i])
    logging.info('end ConvertData')


def main():
    # 解析xls文件到list，用于后续数据处理数据源
    logging.info('welcome to SA Commercial Support world.')
    try:
        XLSReadAndAnalysis()
        ConvertData()
        XLSReadAndWrite()
    except Exception as err:
        logging.error('error: %s', err)

    logging.info("end SA Commercial Support  world")


if __name__ == '__main__':
    main()
